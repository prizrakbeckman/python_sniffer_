import openpyxl, os

def printExcelFile(file):
    if os.path.exists(file):
        wb = openpyxl.load_workbook(file)
        sheets = wb.get_sheet_names()
        print(sheets)
        sheet = wb.get_sheet_by_name(sheets[1])
        i = sheet.max_row
        j = sheet.max_column
        print(i)
        print(j)
        c4 = sheet['C4'].value
        a2 = sheet.cell(row=i, column=j).value
        print('*********'+str(c4)+'************')
        print('*********'+str(a2)+'************')

        #print everything in the sheet
        for rowCellOfObjects in sheet['A1':'C3']:
            for cellObj in rowCellOfObjects:
                print(cellObj.coordinate, cellObj.value)
            print('-- END OF ROW --')

        #print everything in a specific column
        j = input()
        
        while True :
            if j.isdigit():
               if int(j) < sheet.max_column:
                    break
            j = input()
            
            
        print('All good ')
        j = int(j)
       
        l = list(sheet.columns)[j]
        for cellObject in l:
            print(cellObject.value)
        print('End of show of column' + str(j))
        
        sheet.cell(row=5,column=2).value='Pizdets'
        print('Cell updated')
     #   sheet.column_dimensions['B'].hidden = True
        wb.save('test.xlsx')

def createNewWorkBook(file):
    wb = openpyxl.Workbook()
    wb.create_sheet()
    print('Sheet created')
    wb.create_sheet(index=0,title='sheetAdin')
    l = wb.get_sheet_names()
    print(l)
    sheet=wb.get_sheet_by_name('sheetAdin')
    sheet['A1']='Value Added'
    wb.save(file)
        
os.chdir('C:\\Users\\HP\\Desktop\\titiza')
printExcelFile('test.xlsx')
#createNewWorkBook('myTestedWorkbook.xlsx')

    
